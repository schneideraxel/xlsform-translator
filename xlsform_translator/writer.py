"""
Output Excel construction: append translated language columns and save.
"""

import re
import openpyxl

from .parser import ParsedForm, TRANSLATABLE_SHEETS


def _resolve_target_language_string(target_language: str, language_style: str) -> str:
    """
    Produce the canonical language string for the new column headers.

    Mirrors the style of the existing source-language columns:
      language_style "with_code"    → "Spanish (es)"
      language_style "without_code" → "Spanish"

    Uses langcodes for name/code resolution. Falls back to the raw
    user-supplied string if langcodes cannot recognise it.
    """
    try:
        import langcodes
        lang = langcodes.find(target_language)
        name = lang.display_name("en")
        code = lang.language.lower()
        if language_style == "with_code":
            return f"{name} ({code})"
        return name
    except Exception:
        # langcodes couldn't resolve the string — use it as supplied.
        if language_style == "with_code":
            if re.search(r"\([a-z]{2,3}\)$", target_language.strip()):
                return target_language.strip()
            return f"{target_language.strip()} (xx)"
        return re.split(r"\s*\(", target_language)[0].strip()


def _build_target_column_name(source_col_name: str, resolved_target: str, language_style: str) -> str:
    """
    Derive the header for the new target-language column.

    Example: source "label::French (fr)" + target "Spanish (es)" → "label::Spanish (es)"
    """
    base = source_col_name.split("::")[0]
    if language_style == "with_code":
        return f"{base}::{resolved_target}"
    # Strip any trailing code before building the plain-style header.
    name_only = re.split(r"\s*\(", resolved_target)[0].strip()
    return f"{base}::{name_only}"


def build_output(
    parsed: ParsedForm,
    cells: list,
    target_language: str,
    output_path: str,
    verbose: bool = False,
) -> int:
    """
    Write translated columns to the workbook and save it to output_path.

    New columns are always appended at the end of each sheet (sheet.max_column + 1)
    rather than inserted next to their source column. This is intentional:
    openpyxl's insert_cols() shifts cell references but does not update the
    tableColumn metadata inside Excel Table (ListObject) definitions, which
    causes Excel to report the file as corrupt. Since XLSForm platforms
    identify columns by header name rather than position, append-only is safe.

    Args:
        parsed: ParsedForm returned by parse_form().
        cells: The same CellRef list, with .translated_text populated.
        target_language: Target language name or code (used to build headers).
        output_path: Where to write the output .xlsx file.
        verbose: If True, prints skipped-column notices to stdout.

    Returns:
        Number of cells written to the output file.
    """
    resolved_target = _resolve_target_language_string(target_language, parsed.language_style)
    if verbose:
        print(f"  Target language resolved to: {resolved_target}")

    wb = parsed.workbook

    # Build a fast lookup: (sheet_name, row, col_index) → translated text
    translation_lookup = {
        (c.sheet_name, c.row, c.col_index): c.translated_text
        for c in cells
        if hasattr(c, "translated_text")
    }

    cells_written = 0

    for sheet_name in TRANSLATABLE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]

        source_cols = parsed.translatable_columns.get(sheet_name, [])
        source_lang_cols = [c for c in source_cols if c.language == parsed.source_language]
        # Process columns left-to-right to produce a consistent column order.
        source_lang_cols.sort(key=lambda c: c.col_index)

        if not source_lang_cols:
            continue

        existing_headers = [cell.value for cell in sheet[1]]

        for col_info in source_lang_cols:
            new_col_name = _build_target_column_name(
                col_info.name, resolved_target, parsed.language_style
            )
            if new_col_name in existing_headers:
                if verbose:
                    print(f"  Skipping '{new_col_name}' in '{sheet_name}' — already exists.")
                continue

            new_col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=new_col_idx, value=new_col_name)
            existing_headers.append(new_col_name)

            for row_idx in range(2, sheet.max_row + 1):
                translated = translation_lookup.get((sheet_name, row_idx, col_info.col_index))
                if translated is not None:
                    sheet.cell(row=row_idx, column=new_col_idx, value=translated)
                    cells_written += 1

    wb.save(output_path)
    return cells_written
