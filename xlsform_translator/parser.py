### XLSForm parser : column classification, source language matching, and placeholder tokenization
### AS 🐚🫧🪼🪸
### 05.04.2026 (Last update)

"""
XLSForm parsing: column classification, source language matching,
and placeholder tokenization/detokenization.

Parsing happens in three stages:
1. Column classification : identify which columns hold translatable text.
2. Source language matching : find the column language that corresponds to
   what the user specified via --source-language.
3. Cell collection : iterate over all matched columns, tokenize placeholders,
   and return a flat list of CellRef objects ready for translation.
"""

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

# Only these two sheets are processed; all other sheets are ignored.
TRANSLATABLE_SHEETS = ("survey", "choices")

# Column base names that contain user-facing text. Columns whose base name is
# NOT in this set (e.g. bind::*, body::*, media::*) are skipped entirely.
BASE_TRANSLATABLE = {
    "label",
    "hint",
    "guidance_hint",
    "constraint_message",
    "constraint message",   # space variant used by some platforms
    "required_message",
    "required message",     # space variant used by some platforms
}

# Matches content that must not be translated:
#   ${variable}  : XLSForm variable references
#   <tag>        : HTML tags used for rich text formatting
#   #{variable}  : ODK-style calculation references
PLACEHOLDER_RE = re.compile(r"\$\{[^}]+\}|<[^>]+>|#\{[^}]+\}")


@dataclass
class ColumnInfo:
    """Metadata for a single translatable column."""
    name: str        # full column header, e.g. "label::French (fr)"
    base: str        # the translatable base name, e.g. "label"
    language: str    # language portion of the header, e.g. "French (fr)";
                     # empty string ("") for plain columns without a language suffix
    col_index: int   # 1-based column index within the sheet


@dataclass
class CellRef:
    """A reference to a single non-empty translatable cell."""
    sheet_name: str
    row: int           # 1-based row index
    col_index: int     # 1-based column index
    source_text: str   # original cell value
    tokenized_text: str = ""  # source_text with placeholders replaced by [P1], [P2], ...
    token_map: dict = field(default_factory=dict)  # maps "[P1]" -> original placeholder string


@dataclass
class ParsedForm:
    """All data extracted from an XLSForm file, ready for translation."""
    workbook: openpyxl.Workbook
    source_language: str      # exact language string as it appears in column headers
    language_style: str       # "with_code" (e.g. "French (fr)") or "without_code" ("French")
    translatable_columns: dict  # sheet_name -> list[ColumnInfo] for all translatable columns
    cells: list               # flat list[CellRef] of all non-empty source-language cells


# ---------------------------------------------------------------------------
# Column classification
# ---------------------------------------------------------------------------

def _classify_column(name: str, col_index: int) -> Optional[ColumnInfo]:
    """
    Decide whether a column header represents translatable content.

    Rules (applied in order):
    - Exact match to BASE_TRANSLATABLE (e.g. "label") → plain column, language=""
    - "base::language" where base ∈ BASE_TRANSLATABLE (e.g. "label::French (fr)") → language variant
    - Anything else (bind::*, body::*, media::*, calculate, type, name, …) → None (skip)
    """
    if name in BASE_TRANSLATABLE:
        return ColumnInfo(name=name, base=name, language="", col_index=col_index)

    if "::" in name:
        base, _, lang = name.partition("::")
        if base in BASE_TRANSLATABLE:
            return ColumnInfo(name=name, base=base, language=lang, col_index=col_index)

    return None


def classify_columns(sheet) -> list:
    """Return a list of ColumnInfo for every translatable column in a sheet."""
    result = []
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for col_index, cell_value in enumerate(header_row, start=1):
        if cell_value is None:
            continue
        info = _classify_column(str(cell_value).strip(), col_index)
        if info:
            result.append(info)
    return result


def _load_translatable_columns(workbook) -> dict:
    """Build a {sheet_name: [ColumnInfo]} mapping for all TRANSLATABLE_SHEETS."""
    result = {}
    for sheet_name in TRANSLATABLE_SHEETS:
        if sheet_name in workbook.sheetnames:
            cols = classify_columns(workbook[sheet_name])
            if cols:
                result[sheet_name] = cols
    return result


def _col_has_content(sheet, col_index: int) -> bool:
    """Return True if any data row in the column has a non-empty value."""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        val = row[col_index - 1]
        if val and str(val).strip():
            return True
    return False


# ---------------------------------------------------------------------------
# Source language matching
# ---------------------------------------------------------------------------

def _all_language_variants(translatable_columns: dict) -> set:
    """Return the set of all language strings found in column headers."""
    langs = set()
    for cols in translatable_columns.values():
        for col in cols:
            if col.language:
                langs.add(col.language)
    return langs


def _plain_columns_with_content(workbook, translatable_columns: dict) -> list:
    """
    Return the names of plain (no-language-suffix) columns that contain data.
    Used to generate a helpful error message when the form hasn't been
    prepared with language-variant column headers.
    """
    found = []
    for sheet_name in TRANSLATABLE_SHEETS:
        cols = translatable_columns.get(sheet_name, [])
        if not cols:
            continue
        sheet = workbook[sheet_name]
        for col in cols:
            if not col.language and col.name not in found:
                if _col_has_content(sheet, col.col_index):
                    found.append(col.name)
    return found


def _resolve_language_code(language_str: str) -> Optional[str]:
    """
    Resolve any language string format to a two- or three-letter IETF code.

    Handles all common formats that appear in XLSForm column headers:
      "French (fr)"  → "fr"   (extract from parentheses, already explicit)
      "French (FR)"  → "fr"   (uppercase code normalised to lowercase)
      "French"       → "fr"   (name lookup via langcodes)
      "fr"           → "fr"   (bare IETF tag)

    Returns None if the string cannot be resolved (treated as no match).
    """
    # Fast path: code is already embedded in parentheses (case-insensitive).
    m = re.search(r"\(([a-zA-Z]{2,3})\)$", language_str.strip())
    if m:
        return m.group(1).lower()
    try:
        import langcodes
        # Try parsing as a bare IETF tag first ("fr", "sw", "en-US", ...).
        try:
            return langcodes.Language.get(language_str.strip()).language.lower()
        except Exception:
            pass
        # Fall back to name lookup ("French", "Swahili", ...).
        return langcodes.find(language_str.strip()).language.lower()
    except Exception:
        return None


def _match_source_language(translatable_columns: dict, user_language: str) -> Optional[str]:
    """
    Find the language string in the form's column headers that matches the
    user-supplied --source-language value.

    Returns the exact column language string as it appears in the headers
    (e.g. "French (fr)"), or None if no match is found.

    Matching is attempted in two passes:
    1. Case-insensitive exact string match : fastest and most reliable.
    2. IETF code comparison via langcodes : allows "fr" to match "French (fr)",
       or "French" to match "French (FR)".
    """
    available = _all_language_variants(translatable_columns)
    user_lower = user_language.strip().lower()

    # Pass 1: exact case-insensitive match
    for lang in available:
        if lang.strip().lower() == user_lower:
            return lang

    # Pass 2: resolve both sides to IETF codes and compare
    user_code = _resolve_language_code(user_language)
    if user_code:
        for lang in available:
            col_code = _resolve_language_code(lang)
            if col_code and col_code == user_code:
                return lang

    return None


# ---------------------------------------------------------------------------
# Placeholder tokenization
# ---------------------------------------------------------------------------

def tokenize(text: str) -> tuple:
    """
    Replace all placeholders in text with numbered tokens [P1], [P2], ...

    This is done before sending text to a translation engine so that variable
    references, HTML tags, and ODK-style references are protected from being
    translated or mangled. Each unique occurrence gets its own token so that
    repeated references are individually restorable.

    Returns:
        (tokenized_text, token_map) where token_map maps each token string
        (e.g. "[P1]") back to the original placeholder it replaced.
    """
    token_map = {}
    counter = [0]

    def replace(m):
        counter[0] += 1
        token = f"[P{counter[0]}]"
        token_map[token] = m.group(0)
        return token

    tokenized = PLACEHOLDER_RE.sub(replace, text)
    return tokenized, token_map


def detokenize(text: str, token_map: dict) -> str:
    """Restore [P1], [P2], ... tokens to their original placeholder strings."""
    for token, original in token_map.items():
        text = text.replace(token, original)
    return text


# ---------------------------------------------------------------------------
# Top-level parse function
# ---------------------------------------------------------------------------

def parse_form(filepath: str, source_language: str) -> ParsedForm:
    """
    Load an XLSForm workbook and prepare all translatable cells.

    Args:
        filepath: Path to the .xlsx file.
        source_language: Language to translate from. Can be a full name
            ("French"), a name+code ("French (fr)"), or a bare code ("fr").
            Must match a language already present in the column headers.

    Returns:
        A ParsedForm containing the open workbook and a list of CellRef
        objects for every non-empty source-language cell found.

    Raises:
        ValueError: If the form has no language-variant columns, or if
            source_language does not match any language in the form.
    """
    workbook = openpyxl.load_workbook(filepath)
    translatable_columns = _load_translatable_columns(workbook)

    available_languages = _all_language_variants(translatable_columns)
    if not available_languages:
        # Give the user an actionable error: name the plain columns they need to rename.
        plain_with_content = _plain_columns_with_content(workbook, translatable_columns)
        if plain_with_content:
            raise ValueError(
                f"This form has plain columns without a language suffix: "
                f"{', '.join(plain_with_content)}. "
                "Rename your columns to include the language, "
                "e.g. 'label::French (fr)' or 'label::French', "
                "then run the tool again."
            )
        raise ValueError("No translatable content found in this form.")

    matched = _match_source_language(translatable_columns, source_language)
    if not matched:
        raise ValueError(
            f"Source language '{source_language}' not found in this form. "
            f"Available languages: {', '.join(sorted(available_languages))}"
        )

    # Detect whether column headers use the "Language (code)" or "Language" style
    # so that new columns are written in the same format as the existing ones.
    # The regex matches only lowercase codes because matched is already normalised.
    language_style = (
        "with_code"
        if re.search(r"\([a-z]{2,3}\)$", matched)
        else "without_code"
    )

    # Collect all non-empty cells from the matched source-language columns.
    cells = []
    for sheet_name in TRANSLATABLE_SHEETS:
        cols = translatable_columns.get(sheet_name, [])
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        source_cols = [c for c in cols if c.language == matched]
        for col_info in source_cols:
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                val = row[col_info.col_index - 1]
                if val is None:
                    continue
                text = str(val).strip()
                if not text:
                    continue
                tokenized, token_map = tokenize(text)
                cells.append(CellRef(
                    sheet_name=sheet_name,
                    row=row_idx,
                    col_index=col_info.col_index,
                    source_text=text,
                    tokenized_text=tokenized,
                    token_map=token_map,
                ))

    return ParsedForm(
        workbook=workbook,
        source_language=matched,
        language_style=language_style,
        translatable_columns=translatable_columns,
        cells=cells,
    )
